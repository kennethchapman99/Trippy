"""Generate the 5 sample sheet fixtures for testing.

Run once: uv run python tests/fixtures/create_sample_sheets.py
Also called automatically by conftest.py if files are missing.
"""

from __future__ import annotations

from pathlib import Path

FIXTURES_DIR = Path(__file__).parent / "sample_sheets"


def create_column_based_xlsx() -> Path:
    """Standard: headers in row 1, data below — most common format."""
    import openpyxl

    path = FIXTURES_DIR / "column_based.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Japan 2026"  # type: ignore[union-attr]

    # Trip metadata block
    ws.append(["Trip Name", "Japan 2026"])  # type: ignore[union-attr]
    ws.append(["Start Date", "2026-03-15"])  # type: ignore[union-attr]
    ws.append(["End Date", "2026-03-29"])  # type: ignore[union-attr]
    ws.append(["Status", "booked"])  # type: ignore[union-attr]
    ws.append(["Destination", "Tokyo, Kyoto, Osaka"])  # type: ignore[union-attr]
    ws.append([])  # type: ignore[union-attr]

    # Travelers
    ws.append(["Travelers"])  # type: ignore[union-attr]
    ws.append(["Name", "Passport Country", "Passport Expiry"])  # type: ignore[union-attr]
    for row in [
        ("Ken Chapman", "CAN", "2031-06-01"),
        ("Sarah Chapman", "CAN", "2030-11-15"),
        ("Emma Chapman", "CAN", "2029-08-20"),
        ("Liam Chapman", "CAN", "2028-04-10"),
        ("Olivia Chapman", "CAN", "2028-04-10"),
    ]:
        ws.append(list(row))  # type: ignore[union-attr]
    ws.append([])  # type: ignore[union-attr]

    # Flights
    ws.append(["Flights"])  # type: ignore[union-attr]
    ws.append(
        ["Carrier", "Flight #", "From", "To", "Depart", "Arrive", "Class", "Cost CAD", "Conf #"]
    )  # type: ignore[union-attr]
    ws.append(
        [
            "Air Canada",
            "AC003",
            "YYZ",
            "NRT",
            "2026-03-15T13:30:00",
            "2026-03-16T16:45:00",
            "economy",
            "8500",
            "ACJPN26",
        ]
    )  # type: ignore[union-attr]
    ws.append(
        [
            "Air Canada",
            "AC004",
            "NRT",
            "YYZ",
            "2026-03-29T18:00:00",
            "2026-03-29T16:30:00",
            "economy",
            "",
            "ACJPN26R",
        ]
    )  # type: ignore[union-attr]
    ws.append([])  # type: ignore[union-attr]

    # Hotels
    ws.append(["Hotels"])  # type: ignore[union-attr]
    ws.append(["Property", "City", "Country", "Check-in", "Check-out", "Cost CAD", "Conf #"])  # type: ignore[union-attr]
    ws.append(
        [
            "Shinjuku Granbell Hotel",
            "Tokyo",
            "Japan",
            "2026-03-16",
            "2026-03-22",
            "2800",
            "SGH-4421",
        ]
    )  # type: ignore[union-attr]
    ws.append(
        [
            "The Westin Miyako Kyoto",
            "Kyoto",
            "Japan",
            "2026-03-22",
            "2026-03-25",
            "2100",
            "WMK-8832",
        ]
    )  # type: ignore[union-attr]

    wb.save(path)
    return path


def create_row_based_csv() -> Path:
    """Row-based: labels in column A, values in column B."""
    path = FIXTURES_DIR / "row_based.csv"
    path.write_text(
        """\
Field,Value
Trip Name,Costa Rica Winter 2025
Start Date,2025-12-20
End Date,2025-12-31
Status,lived
Destination,San José / La Fortuna / Manuel Antonio
,
Travelers,
Name,Passport
Ken Chapman,CAN
Sarah Chapman,CAN
Emma Chapman,CAN
Liam Chapman,CAN
Olivia Chapman,CAN
,
Flight,
Carrier,United Airlines
Flight Number,UA1842
From,YYZ
To,SJO
Departure,2025-12-20T07:15:00
Arrival,2025-12-20T14:30:00
Cabin,economy
Cost CAD,6200
Confirmation,UA-CR2025
,
Accommodation,
Property,Jungle Villa La Fortuna
Type,airbnb
City,La Fortuna
Country,Costa Rica
Check In,2025-12-21
Check Out,2025-12-26
Cost CAD,1900
Confirmation,HMTA9BX2
""",
        encoding="utf-8",
    )
    return path


def create_multi_tab_xlsx() -> Path:
    """Multi-tab: separate sheets for trip info, flights, and hotels."""
    import openpyxl

    path = FIXTURES_DIR / "multi_tab.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1: Trip overview
    ws_trip = wb.active
    ws_trip.title = "Trip Overview"  # type: ignore[union-attr]
    ws_trip.append(["Trip Name", "Portugal 2027"])  # type: ignore[union-attr]
    ws_trip.append(["Start", "2027-06-10"])  # type: ignore[union-attr]
    ws_trip.append(["End", "2027-06-24"])  # type: ignore[union-attr]
    ws_trip.append(["Status", "planned"])  # type: ignore[union-attr]
    ws_trip.append(["Destinations", "Lisbon, Porto, Douro Valley"])  # type: ignore[union-attr]
    ws_trip.append(["Travelers", "Ken, Sarah, Emma, Liam, Olivia"])  # type: ignore[union-attr]

    # Sheet 2: Flights
    ws_flights = wb.create_sheet("Flights")
    ws_flights.append(["Carrier", "Flight", "From", "To", "Depart", "Arrive", "Cost"])
    ws_flights.append(
        [
            "TAP Air Portugal",
            "TP253",
            "YYZ",
            "LIS",
            "2027-06-10T19:45:00",
            "2027-06-11T07:30:00",
            "9500",
        ]
    )
    ws_flights.append(
        [
            "TAP Air Portugal",
            "TP254",
            "LIS",
            "YYZ",
            "2027-06-24T10:00:00",
            "2027-06-24T13:15:00",
            "",
        ]
    )

    # Sheet 3: Accommodation
    ws_hotels = wb.create_sheet("Hotels")
    ws_hotels.append(["Property", "Type", "City", "Check-in", "Check-out", "Cost CAD"])
    ws_hotels.append(["Bairro Alto Hotel", "hotel", "Lisbon", "2027-06-11", "2027-06-17", "4200"])
    ws_hotels.append(["Yeatman Hotel", "hotel", "Porto", "2027-06-17", "2027-06-21", "3100"])

    wb.save(path)
    return path


def create_free_form_csv() -> Path:
    """Free-form: mixed prose and structured data, no consistent headers."""
    path = FIXTURES_DIR / "free_form.csv"
    path.write_text(
        """\
CHAPMAN FAMILY - ICELAND ADVENTURE
July 2027 trip - Reykjavik and the ring road - 12 nights

FLIGHTS
We're flying Air Canada YYZ → KEF on July 5 2027 (AC849) departing 6pm arriving July 6 6am
Return July 17 2027 KEF → YYZ on AC850 at 10:30am
Estimated cost for 5 pax: ~CAD 11000 total

ACCOMMODATION
- July 6-10: Ion Adventure Hotel near Selfoss (hotel) check out July 10 - cost $3200 CAD conf IONADV77
- July 10-13: Husafell Mountain Bungalows (house) Husafell village - $1800 CAD
- July 13-17: Reykjavik Lights Hotel downtown (hotel) - $2600 CAD conf RLH-5519

TRAVELERS (all Canadian passports)
Ken - expires June 2031
Sarah - expires Nov 2030
3 kids (Emma Liam Olivia) - passports expire April 2028
""",
        encoding="utf-8",
    )
    return path


def create_broken_csv() -> Path:
    """Broken: malformed/incomplete data — importer must handle gracefully."""
    path = FIXTURES_DIR / "broken.csv"
    content = (
        "Trip,,,,\n"
        ",2026\n"
        "Travelers,,,,\n"
        "Name,,,\n"
        '"Ken, ""the traveler""",bad date,???\n'
        ",,,missing fields\n"
        "Flights\n"
        "From,To\n"
        "YYZ\n"
        ",,,,,,,,,,,,,,\n"
        "Status,UNKNOWN_STATUS_VALUE\n"
        "Cost,not-a-number\n"
        "Date,32/13/2026\n"
    )
    path.write_text(content, encoding="utf-8")
    return path


def create_all(target_dir: Path | None = None) -> dict[str, Path]:
    global FIXTURES_DIR
    if target_dir is not None:
        FIXTURES_DIR = target_dir
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)

    return {
        "column_based": create_column_based_xlsx(),
        "row_based": create_row_based_csv(),
        "multi_tab": create_multi_tab_xlsx(),
        "free_form": create_free_form_csv(),
        "broken": create_broken_csv(),
    }


if __name__ == "__main__":
    paths = create_all()
    for name, path in paths.items():
        print(f"Created {name}: {path}")
