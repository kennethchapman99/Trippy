"""Sheet importer — normalises xlsx/csv/Google Sheets into Trip records via Claude."""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import anthropic
from pydantic import BaseModel
from sqlalchemy.orm import Session

from hermes_trip import config
from hermes_trip.db import make_session_factory
from hermes_trip.db.models import Leg, LegType, Stay, StayType, Traveler, Trip, TripStatus

logger = logging.getLogger(__name__)

CONFIDENCE_THRESHOLD = 0.7
MAX_ROWS_PER_SHEET = 300

# ---------------------------------------------------------------------------
# Pydantic models for Claude's structured output
# ---------------------------------------------------------------------------


class FieldValue(BaseModel):
    value: str | None = None
    confidence: float = 0.0


class ParsedTraveler(BaseModel):
    name: FieldValue = FieldValue()
    passport_country: FieldValue = FieldValue()
    passport_expiry: FieldValue = FieldValue()


class ParsedLeg(BaseModel):
    leg_type: FieldValue = FieldValue(value="flight", confidence=0.8)
    carrier: FieldValue = FieldValue()
    flight_number: FieldValue = FieldValue()
    origin: FieldValue = FieldValue()
    destination: FieldValue = FieldValue()
    depart_at: FieldValue = FieldValue()
    arrive_at: FieldValue = FieldValue()
    cabin_class: FieldValue = FieldValue()
    cost_cad: FieldValue = FieldValue()
    confirmation_code: FieldValue = FieldValue()


class ParsedStay(BaseModel):
    stay_type: FieldValue = FieldValue(value="hotel", confidence=0.8)
    property_name: FieldValue = FieldValue()
    city: FieldValue = FieldValue()
    country: FieldValue = FieldValue()
    check_in: FieldValue = FieldValue()
    check_out: FieldValue = FieldValue()
    cost_cad: FieldValue = FieldValue()
    confirmation_code: FieldValue = FieldValue()


class ParsedTrip(BaseModel):
    name: FieldValue = FieldValue()
    start_date: FieldValue = FieldValue()
    end_date: FieldValue = FieldValue()
    status: FieldValue = FieldValue(value="planned", confidence=0.5)
    destination_summary: FieldValue = FieldValue()
    travelers: list[ParsedTraveler] = []
    legs: list[ParsedLeg] = []
    stays: list[ParsedStay] = []


class ParseResult(BaseModel):
    trips: list[ParsedTrip] = []
    parsing_notes: str = ""


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------


@dataclass
class FlaggedField:
    trip_name: str
    section: str  # "trip" | "leg:N" | "stay:N" | "traveler:N"
    field: str
    value: str | None
    confidence: float


@dataclass
class ImportResult:
    source: str = ""
    trips_created: int = 0
    trips_updated: int = 0
    flagged_fields: list[FlaggedField] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    parsing_notes: str = ""

    @property
    def ok(self) -> bool:
        return not self.errors


# ---------------------------------------------------------------------------
# Claude tool schema
# ---------------------------------------------------------------------------

_FIELD_STR = {
    "type": "object",
    "required": ["value", "confidence"],
    "properties": {
        "value": {"type": ["string", "null"]},
        "confidence": {"type": "number", "minimum": 0.0, "maximum": 1.0},
    },
}

_TRAVELER_SCHEMA = {
    "type": "object",
    "properties": {
        "name": _FIELD_STR,
        "passport_country": _FIELD_STR,
        "passport_expiry": _FIELD_STR,
    },
}

_LEG_SCHEMA = {
    "type": "object",
    "properties": {
        "leg_type": _FIELD_STR,
        "carrier": _FIELD_STR,
        "flight_number": _FIELD_STR,
        "origin": _FIELD_STR,
        "destination": _FIELD_STR,
        "depart_at": _FIELD_STR,
        "arrive_at": _FIELD_STR,
        "cabin_class": _FIELD_STR,
        "cost_cad": _FIELD_STR,
        "confirmation_code": _FIELD_STR,
    },
}

_STAY_SCHEMA = {
    "type": "object",
    "properties": {
        "stay_type": _FIELD_STR,
        "property_name": _FIELD_STR,
        "city": _FIELD_STR,
        "country": _FIELD_STR,
        "check_in": _FIELD_STR,
        "check_out": _FIELD_STR,
        "cost_cad": _FIELD_STR,
        "confirmation_code": _FIELD_STR,
    },
}

EXTRACT_TRIPS_TOOL: dict[str, Any] = {
    "name": "extract_trips",
    "description": (
        "Extract all trip records from the given spreadsheet content. "
        "For each field provide a confidence score: 1.0=explicit, 0.8=clear inference, "
        "0.6=reasonable guess, 0.4=uncertain, 0.0=absent/unknown. "
        "Dates: ISO 8601. Airport codes: IATA 3-letter. Passport country: ISO 3166-1 alpha-3. "
        "Costs: numeric string in CAD."
    ),
    "input_schema": {
        "type": "object",
        "required": ["trips", "parsing_notes"],
        "properties": {
            "trips": {
                "type": "array",
                "items": {
                    "type": "object",
                    "required": ["name", "start_date"],
                    "properties": {
                        "name": _FIELD_STR,
                        "start_date": _FIELD_STR,
                        "end_date": _FIELD_STR,
                        "status": _FIELD_STR,
                        "destination_summary": _FIELD_STR,
                        "travelers": {"type": "array", "items": _TRAVELER_SCHEMA},
                        "legs": {"type": "array", "items": _LEG_SCHEMA},
                        "stays": {"type": "array", "items": _STAY_SCHEMA},
                    },
                },
            },
            "parsing_notes": {"type": "string"},
        },
    },
}

SYSTEM_PROMPT = """You are a travel data extraction specialist working with the Chapman family \
(5 travelers, Oakville Ontario, Canada). Extract ALL trips from the spreadsheet.

Rules:
- Dates: ISO 8601 (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS)
- Airport codes: IATA 3-letter uppercase (YYZ, NRT, LHR)
- Passport country: ISO 3166-1 alpha-3 (CAN, USA, JPN, GBR)
- Costs: numeric string in CAD — convert if another currency is labelled (lower confidence)
- Leg types: flight, train, ferry, bus, car, other
- Stay types: hotel, airbnb, vrbo, hostel, house, other
- Trip status: dream, planned, booked, lived, cancelled
- Set value=null and confidence=0.0 for missing/unknown fields
- If the sheet is broken or empty, return trips=[] with a parsing_notes explanation"""


# ---------------------------------------------------------------------------
# Sheet reader helpers
# ---------------------------------------------------------------------------


def _read_xlsx(path: Path) -> str:
    """Convert xlsx to a plain-text table representation."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_ROWS_PER_SHEET:
                rows.append(f"... (truncated at {MAX_ROWS_PER_SHEET} rows)")
                break
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):  # skip fully empty rows
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def _read_csv(path: Path) -> str:
    """Read CSV file to plain text, trying common encodings."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = path.read_text(encoding=enc)
            # Normalise to TSV-ish format for consistency
            reader = csv.reader(io.StringIO(text))
            rows = []
            for i, row in enumerate(reader):
                if i >= MAX_ROWS_PER_SHEET:
                    rows.append(f"... (truncated at {MAX_ROWS_PER_SHEET} rows)")
                    break
                rows.append("\t".join(row))
            return "\n".join(rows)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not decode {path} with utf-8-sig, utf-8, or latin-1")


def _read_google_sheets(url: str) -> str:
    """Export Google Sheets URL to CSV text (requires public sharing or API key)."""
    import urllib.request

    # Convert edit/view URL to CSV export URL
    csv_url = re.sub(r"/edit.*$", "/export?format=csv", url)
    csv_url = re.sub(r"/view.*$", "/export?format=csv", csv_url)
    if "export" not in csv_url:
        csv_url = url.rstrip("/") + "/export?format=csv"
    try:
        with urllib.request.urlopen(csv_url, timeout=30) as resp:  # noqa: S310
            raw = resp.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(raw))
        return "\n".join("\t".join(row) for row in reader)
    except Exception as exc:
        raise ValueError(
            f"Could not fetch Google Sheets URL. "
            f"Make sure the sheet is publicly shared (Anyone with link → Viewer). "
            f"Error: {exc}"
        ) from exc


def read_sheet_to_text(source: str | Path) -> str:
    """Read any supported source (path or URL) into plain text for Claude."""
    s = str(source)
    if s.startswith("https://docs.google.com/spreadsheets"):
        return _read_google_sheets(s)
    path = Path(s)
    if not path.exists():
        raise FileNotFoundError(f"Sheet file not found: {path}")
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return _read_xlsx(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise ValueError(f"Unsupported file type: {suffix}. Supported: .xlsx, .csv")


# ---------------------------------------------------------------------------
# Date / value helpers
# ---------------------------------------------------------------------------


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _parse_float(value: str | None) -> float | None:
    if not value:
        return None
    cleaned = re.sub(r"[^\d.\-]", "", value)
    try:
        return float(cleaned)
    except ValueError:
        return None


def _safe_enum(cls: type, value: str | None, default: Any) -> Any:
    if not value:
        return default
    try:
        return cls(value.lower().strip())
    except ValueError:
        return default


# ---------------------------------------------------------------------------
# Flag collection
# ---------------------------------------------------------------------------


def _flag_field(
    flags: list[FlaggedField],
    trip_name: str,
    section: str,
    field_name: str,
    fv: FieldValue,
) -> None:
    if fv.confidence < CONFIDENCE_THRESHOLD:
        flags.append(
            FlaggedField(
                trip_name=trip_name,
                section=section,
                field=field_name,
                value=fv.value,
                confidence=fv.confidence,
            )
        )


def _collect_flags(parsed: ParsedTrip) -> list[FlaggedField]:
    name = parsed.name.value or "unknown"
    flags: list[FlaggedField] = []

    for f_name in ("name", "start_date", "end_date", "status", "destination_summary"):
        _flag_field(flags, name, "trip", f_name, getattr(parsed, f_name))

    for i, t in enumerate(parsed.travelers):
        for f_name in ("name", "passport_country", "passport_expiry"):
            _flag_field(flags, name, f"traveler:{i}", f_name, getattr(t, f_name))

    for i, leg in enumerate(parsed.legs):
        for f_name in (
            "leg_type",
            "carrier",
            "origin",
            "destination",
            "depart_at",
            "arrive_at",
            "cost_cad",
        ):
            _flag_field(flags, name, f"leg:{i}", f_name, getattr(leg, f_name))

    for i, stay in enumerate(parsed.stays):
        for f_name in ("property_name", "city", "country", "check_in", "check_out"):
            _flag_field(flags, name, f"stay:{i}", f_name, getattr(stay, f_name))

    return flags


# ---------------------------------------------------------------------------
# DB upsert
# ---------------------------------------------------------------------------


def _upsert_trip(session: Session, parsed: ParsedTrip) -> tuple[Trip, bool]:
    """Insert or update Trip + children. Returns (trip, created)."""
    trip_name = parsed.name.value or "Unnamed Trip"
    start = _parse_date(parsed.start_date.value) or date.today()

    existing = session.query(Trip).filter_by(name=trip_name, start_date=start).first()
    created = existing is None
    trip: Trip = existing or Trip()

    trip.name = trip_name
    trip.start_date = start
    trip.end_date = _parse_date(parsed.end_date.value)
    trip.status = _safe_enum(TripStatus, parsed.status.value, TripStatus.planned)
    trip.destination_summary = parsed.destination_summary.value

    if created:
        session.add(trip)
        session.flush()
    else:
        # Remove old children so we can re-insert from the latest parse
        for child_list in (trip.travelers, trip.legs, trip.stays):
            for child in list(child_list):
                session.delete(child)
        session.flush()

    for t in parsed.travelers:
        if not t.name.value:
            continue
        session.add(
            Traveler(
                trip_id=trip.id,
                name=t.name.value,
                passport_country=t.passport_country.value,
                passport_expiry=_parse_date(t.passport_expiry.value),
            )
        )

    for leg in parsed.legs:
        if not leg.origin.value or not leg.destination.value:
            continue
        session.add(
            Leg(
                trip_id=trip.id,
                leg_type=_safe_enum(LegType, leg.leg_type.value, LegType.flight),
                carrier=leg.carrier.value,
                flight_number=leg.flight_number.value,
                origin=leg.origin.value,
                destination=leg.destination.value,
                depart_at=_parse_datetime(leg.depart_at.value),
                arrive_at=_parse_datetime(leg.arrive_at.value),
                cabin_class=leg.cabin_class.value,
                cost_cad=_parse_float(leg.cost_cad.value),
                confirmation_code=leg.confirmation_code.value,
            )
        )

    for stay in parsed.stays:
        if not stay.property_name.value:
            continue
        session.add(
            Stay(
                trip_id=trip.id,
                stay_type=_safe_enum(StayType, stay.stay_type.value, StayType.hotel),
                property_name=stay.property_name.value,
                city=stay.city.value,
                country=stay.country.value,
                check_in=_parse_date(stay.check_in.value),
                check_out=_parse_date(stay.check_out.value),
                cost_cad=_parse_float(stay.cost_cad.value),
                confirmation_code=stay.confirmation_code.value,
            )
        )

    session.flush()
    return trip, created


# ---------------------------------------------------------------------------
# Main importer class
# ---------------------------------------------------------------------------


class SheetImporter:
    def __init__(
        self,
        db_url: str | None = None,
        anthropic_client: anthropic.Anthropic | None = None,
    ) -> None:
        self._db_url = db_url
        self._client = anthropic_client or anthropic.Anthropic(api_key=config.ANTHROPIC_API_KEY)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def import_file(self, source: str | Path) -> ImportResult:
        """Import a single xlsx/csv file or Google Sheets URL."""
        result = ImportResult(source=str(source))
        try:
            sheet_text = read_sheet_to_text(source)
        except Exception as exc:
            result.errors.append(f"Could not read file: {exc}")
            return result

        try:
            parse_result = self._call_claude(sheet_text)
        except Exception as exc:
            result.errors.append(f"Claude API error: {exc}")
            return result

        result.parsing_notes = parse_result.parsing_notes

        factory = make_session_factory(self._db_url)
        with factory() as session:
            for parsed_trip in parse_result.trips:
                if not parsed_trip.name.value:
                    result.errors.append("Skipped a trip with no name")
                    continue
                try:
                    _, created = _upsert_trip(session, parsed_trip)
                    if created:
                        result.trips_created += 1
                    else:
                        result.trips_updated += 1
                    result.flagged_fields.extend(_collect_flags(parsed_trip))
                except Exception as exc:
                    session.rollback()
                    result.errors.append(f"DB error for trip '{parsed_trip.name.value}': {exc}")
            session.commit()

        return result

    def import_folder(self, folder: Path, dry_run: bool = False) -> list[ImportResult]:
        """Import all xlsx/csv files in a directory."""
        paths = sorted(
            p for p in folder.iterdir() if p.suffix.lower() in (".xlsx", ".xlsm", ".csv")
        )
        results = []
        for path in paths:
            if dry_run:
                try:
                    text = read_sheet_to_text(path)
                    r = ImportResult(source=str(path))
                    r.parsing_notes = f"[dry-run] {len(text.splitlines())} lines readable"
                    results.append(r)
                except Exception as exc:
                    r = ImportResult(source=str(path))
                    r.errors.append(str(exc))
                    results.append(r)
            else:
                results.append(self.import_file(path))
        return results

    # ------------------------------------------------------------------
    # Claude call
    # ------------------------------------------------------------------

    def _call_claude(self, sheet_text: str) -> ParseResult:
        from anthropic.types import ToolChoiceToolParam, ToolParam

        tool: ToolParam = {
            "name": EXTRACT_TRIPS_TOOL["name"],
            "description": EXTRACT_TRIPS_TOOL["description"],
            "input_schema": EXTRACT_TRIPS_TOOL["input_schema"],
        }
        tool_choice: ToolChoiceToolParam = {"type": "tool", "name": "extract_trips"}

        message = self._client.messages.create(
            model="claude-opus-4-7",
            max_tokens=4096,
            system=SYSTEM_PROMPT,
            tools=[tool],
            tool_choice=tool_choice,
            messages=[
                {
                    "role": "user",
                    "content": (
                        f"Please extract all trip data from this spreadsheet:\n\n{sheet_text}"
                    ),
                }
            ],
        )
        # Find the tool_use block; use getattr so MagicMock works in tests too
        from anthropic.types import ToolUseBlock

        for block in message.content:
            if isinstance(block, ToolUseBlock) and block.name == "extract_trips":
                return ParseResult.model_validate(block.input)
            # Fallback path for MagicMock in tests (block.type is a string attribute on mock)
            b_type = getattr(block, "type", None)
            b_name = getattr(block, "name", None)
            b_input = getattr(block, "input", None)
            if b_type == "tool_use" and b_name == "extract_trips" and b_input is not None:
                return ParseResult.model_validate(b_input)
        raise RuntimeError("Claude did not return an extract_trips tool call")
